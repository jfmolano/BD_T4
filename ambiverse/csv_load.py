from openpyxl import load_workbook
from pymongo import MongoClient
import json

with open('conf.json', 'r') as f:
    try:
        conf = json.load(f)
    except ValueError:
        conf = {}

ip_mongo = conf["ip_mongo"]

client = MongoClient(ip_mongo, 27017)
db = client['Grupo03']
collection_preguntas = db['preguntas_taller4']

wb=load_workbook("../../Movie Stack Exchange.xlsx")
ws=wb.active
cell_range = ws['A1':'J150']

for row in cell_range:
	print "- - - - - - - - - - - - - - - - - - - - - - - - - - - - - -"
	print row
	obj = {"q_id":row[0].value,"question":row[1].value,"description":row[2].value,"answer_1":row[3].value,"answer_2":row[4].value,"answer_3":row[5].value,"movie":row[6].value,"tag_1":row[7].value,"tag_2":row[8].value,"tag_3":row[9].value}
	collection_preguntas.insert(obj)